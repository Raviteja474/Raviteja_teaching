import json
import random
import time

import openpyxl

# import this
import constants

# load/read the student files(.json) and feed it to dictionaries.
# assign the application number using random module
# attend for exam and check Adhar card
# give result after 10 secs, let them enjoy holidays
# give ranks
# ask them for counselling
# give seat allotment
# ask their parent to pay fee
# store payment details(text_file), student info(JSON file), total info (excel file)
# text file = payment receipt
# json file = student details
# excel file = college seat allotment details.

print("EAMCET NOTIFICATION************************************")
# with context manager you don't need to explicitely close the file, used for memory management.
# open method open any file txt/json/xlsx
# read mode
# as alias name or file object


# serialization/marshalling = dict-> json(dump)
# de serialization/unmarshalling = json-> dict(load)

with open(constants.EAMCET_NOTICATION_URL, "r")as f:
    # file_object.readlines with create list of contents in total file
    lines = f.readlines()
    # printing line by line
    for line in lines:
        print(line)

print("Adding application number to students.")

# common prefix will be there for all students for application number
APPLICATION = "APEMCT2021"

# setting up default string for set available info
seats_available_info = {"ECE":True, "EEE":True, "CSE":True, "IT":True,"MECH":True, "CIVIL":True}


# load student info and add application number, default file mode read
def get_student_json_info(file_location,mode="r"):
    with open(file_location, "r") as j:
        # we don't need to load text file but as json having a particular data structure i,e dictionary
        # load method will convert string of values from .json file to dictionary in python file
        student_info = json.load(j)
        # 19@@ giving application number ending with 4 random numbers
        application_id = APPLICATION + str(random.randrange(1000,9999))
        student_info["Application ID"]= application_id
        return student_info



# 19@@
def give_student_rank(student_info):
    # alredy_assign_rank = []
    # rank_random = random.choice(RANKS)
    # while rank_random not in alredy_assign_rank:
    #     alredy_assign_rank.append(rank_random)
    #     rank_random = random.choice(RANKS)
    #     break
    # print(f"rank assigned is {rank_random} to ")
    # creating new item with key EAMCET_rank and value rank as taken randonly from list of ranks 1-6
    # assigning EAMCET ranks randomly
    rank_random = random.randrange(1000,9999)
    student_info["EAMCET_rank"] = rank_random
    return student_info


def get_preference_list(student_info):
    branches = ["ECE", "EEE", "CSE", "IT", "MECH", "CIVIL"]
    # Each student will have his own taste of branches. that's why we are shuffling so that everyone will have
    # different preference list
    random.shuffle(branches)
    # updating student info dictionary with preference_list
    student_info["prefernce_list"] = branches
    return student_info


def get_branch_allocation(studnet_info):

   print(f"Student: {studnet_info['Name']} with prefernce list : {studnet_info['prefernce_list']}")
    # searching branch in preference list of candidate.
   for branch in studnet_info['prefernce_list']:
       # if still seat available , allocate to this candidate.
       if seats_available_info[branch]:
           # adding branch_joined key and with value of the "branch" to dictionaty
           studnet_info["branch_joined"] = branch
           # Make unavaialable this branch to next student as it is taken by this student.
           seats_available_info[branch] = False
           # student allocated branch, no need to search , so break it.
           break


def fee_payment(student_info):

    # Preserve payment receipt text file and create json file for the student once his father pays the fee.
    branch_joined = student_info["branch_joined"]
    fees = branch_fees[branch_joined]
    print(f"Mr. {student_info['Father_name']},  please pay the fees amount of {fees}.")
    # please see below complex code as well.
    # print(f"Mr. {student_info['Father_name']},  please pay the fees amount of {branch_fees[student_info["branch_joined"]]}.")

    print("Amount paid. Please wait until we generate online receipt")

    # with open(constants.FEES_INFO_URL, encoding= "utf16") as f:
    #     lines = f.readlines()
    # 19@@
    # UnicodeDecodeError: 'utf-16-le' codec can't decode bytes in position 618-619: illegal encoding

    # Forming the complex string as we are geeting error , so with concatenation of strings we formed resultant(big) string.

    content = 'student: '+ student_info['Name']+' son of: '+ student_info['Father_name'] +' bearing adhar card number: '\
              + student_info['Adhar']+' and paid fees: '+str(fees)+'\n'+ "Thanks for paying fees.\nYours \nCollege Management"
    # Forming text file name in records folder with student name with .txt extension.
    file_name = constants.STORE_RECORDS+student_info['Name']+'.txt'
    # creating text file in write mode and writing the content.
    with open(file_name, 'w') as f:
        f.write(content)


def preserve_records(student_info):
    print(f"Stroring information of {student_info['Name']}.")
    # # Forming text file name in records folder with student name with .json extension.
    file_name = constants.STORE_RECORDS+student_info['Name']+'.json'
    with open(file_name, 'w') as f:
        # first argument is a dictionary which we need to store.
        # second argument is a file object in which we will store data.
        # third argument indent will tell show up space we need to give.
        json.dump(student_info, f, indent=4)


# passing json file location and opening it in read mode in default method argument
# 19@@ write better code.
avinash_info = get_student_json_info(constants.AVINASH_URL)
print(avinash_info)
bharath_info = get_student_json_info(constants.BHARATH_URL)
print(bharath_info)
mahesh_info = get_student_json_info(constants.MAHESH_URL)
print(mahesh_info)
raviteja_info = get_student_json_info(constants.RAVITEJA_URL)
print(raviteja_info)
vishnu_info = get_student_json_info(constants.VISHNU_URL)
print(vishnu_info)
visweswar_info = get_student_json_info(constants.VISWESWAR_URL)
print(visweswar_info)

# multithreading 19@@ check adhar in simultaneous.
print("_____________________________________________Welcome to EAMCET-2021 and please wear mask and maintain social"
      " distance______________________________________________")
print("DON'T COPY :)")
time.sleep(5)
print("______________________________________________Thanks for attending exam. We will announce results shortly."
      "_________________________________________________________")


starting_time = time.time()
while True:
    if time.time()-starting_time < 3:
        # pickup a random ball every time
        print("Enjoying holidays :):):):):):):):):):):):):):):):):):):):):):):):):):):):)")
        time.sleep(1)
    else:
        break

print("_____________________________________Holidays over. Your Ranks are out, plesse check in internet."
      "________________________________")
RANKS = [1, 2, 3, 4, 5, 6]
# giving 6 students , 6 ranks 19@@ , don't give duplicate ranks
avinash_info = give_student_rank(avinash_info)
print(avinash_info)
bharath_info = give_student_rank(bharath_info)
print(bharath_info)
mahesh_info = give_student_rank(mahesh_info)
print(mahesh_info)
raviteja_info = give_student_rank(raviteja_info)
print(raviteja_info)
vishnu_info = give_student_rank(vishnu_info)
print(vishnu_info)
visweswar_info = give_student_rank(visweswar_info)
print(visweswar_info)


print("_________________________________________Congratulations all for qualifying EAMCET , please attend counselling."
      "__________________________________________")

# Avi Civil, ECE, EEE  6000
# visu CSE, EEE, IT   5000
# Ravi CSE , IT, Civil  7000

#  prepare student branch preference list
avinash_info = get_preference_list(avinash_info)
print(avinash_info)
bharath_info = get_preference_list(bharath_info)
print(bharath_info)
mahesh_info = get_preference_list(mahesh_info)
print(mahesh_info)
raviteja_info = get_preference_list(raviteja_info)
print(raviteja_info)
vishnu_info = get_preference_list(vishnu_info)
print(vishnu_info)
visweswar_info = get_preference_list(visweswar_info)
print(visweswar_info)

print("_____________________________________Seat counselling starts.________________________________")
# storing all branches to list from an excel file
branches = []
wb = openpyxl.load_workbook(constants.FEES_INFO_URL)
# finding the active sheet from excel file
sheet_obj = wb.active
# reading the cell
for row_value in range(2, 8):
    # branches are in column 1
    cell_obj = sheet_obj.cell(row=row_value, column=1)
    # printing the cell value
    print("Adding ",cell_obj.value)
    branches.append(cell_obj.value)

print(f"All available branches in GTNN {branches}")

print("__________________________________FEE information____________________________________________________")
# storing all branches and fees info to dictionary from an excel file
branch_fees = {}
wb = openpyxl.load_workbook(constants.FEES_INFO_URL)
# finding the active sheet from excel file
sheet_obj = wb.active
# reading the cell
for row_value in range(2, 8):
    # branches are in column 1
    cell_branch = sheet_obj.cell(row=row_value, column=1)
    # branches are in column 1
    print("cell_branch ",cell_branch.value)
    # fees details are in column 2
    cell_fees = sheet_obj.cell(row=row_value, column=2)
    print("cell_fees ",cell_fees.value)
    branch_fees[cell_branch.value] = cell_fees.value

print("Printing branch fee information:",branch_fees)

print("______________________________________sorting students based upon rank__________________________")

ranks_list = []
# iterating over multiple student and storing their ranks in lis
for student in [avinash_info, bharath_info, mahesh_info, raviteja_info, vishnu_info, visweswar_info]:
    ranks_list.append(student['EAMCET_rank'])

print(f"Unsorted ranks are: {ranks_list}")
ranks_list.sort()
# sort the list in ascending order. Lower rank should go to counselling first followed by higher rank studnets.
print(f"Sorted ranks are: {ranks_list}")

sorted_student_info = []

# we are trying to create list with student info like below
# [rank 1 student dictionary, rank 2 student dictionary, rank 3 student dictionary, rank 4 student dictionary,
# rank 5 student dictionary, rank 6 student dictionary]

# here ranks_list in ascending order
for rank in ranks_list:
    for student_info in [avinash_info, bharath_info, mahesh_info, raviteja_info, vishnu_info, visweswar_info]:
        # adding to lower rank student dictionary as a element to the list sorted_student_info
        if rank == student_info['EAMCET_rank']:
            sorted_student_info.append(student_info)

print(f"Sorted students based on rank: {sorted_student_info}")

print("____________________________seat counselling___________________________________________________________")

for student_info in sorted_student_info:
    # This method will allocate branches based on rank
    get_branch_allocation(student_info)
    print(student_info)

print("____________________________seat counselling___________________________________________________________")

for student_info in sorted_student_info:
    # This method will store fee payment receipt
    fee_payment(student_info)

for student_info in sorted_student_info:
    # This method will store total student information in json file
    preserve_records(student_info)


